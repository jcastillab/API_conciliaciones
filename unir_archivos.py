import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font 
from openpyxl.utils import get_column_letter
from io import BytesIO

def conciliar_movimientos(contabilidad_path: str, df_extracto: pd.DataFrame) -> str:    
    

    df1 = contabilidad_path.copy()
    df2 = df_extracto.copy()   

    # --- Preparación del df1 ---
    df1['FECHA'] = pd.to_datetime(df1['FECHA']).dt.strftime('%Y-%m-%d')
    df1['VALOR'] = df1['VALOR'].astype(int)
    df1['clave_unica'] = df1['FECHA'] + '_' + df1['VALOR'].astype(str)

    # --- Preparación del df2 ---
    df2['VALOR'] = df2['VALOR'].astype(int)
    df2['FECHA'] = pd.to_datetime(df2['FECHA']).dt.strftime('%Y-%m-%d')
    df2['clave_unica'] = df2['FECHA'] + '_' + df2['VALOR'].astype(str) 



    merged_df = pd.merge(df1, df2, on='clave_unica', how='outer', suffixes=('_Contabilidad', '_Extracto'))
    consolidado = merged_df.copy()

    # Caso 1: Entradas en contabilidad y no en extracto
    caso_1 = merged_df[(merged_df['VALOR_Contabilidad'] > 0) & (merged_df['VALOR_Extracto'].isna())]
    caso_1 = caso_1.drop(['FECHA_Extracto', 'VALOR_Extracto','clave_unica','DESCRIPCION'], axis=1)
    # Caso 2: Entradas en extracto y no en contabilidad
    caso_2 = merged_df[(merged_df['VALOR_Extracto'] > 0) & (merged_df['VALOR_Contabilidad'].isna())]
    caso_2 = caso_2.drop(['FECHA_Contabilidad', 'VALOR_Contabilidad','Concepto Contabilidad','clave_unica'], axis=1)
    # Caso 3: Salidas en contabilidad y no en extracto
    caso_3 = merged_df[(merged_df['VALOR_Contabilidad'] < 0) & (merged_df['VALOR_Extracto'].isna())]
    caso_3 = caso_3.drop(['FECHA_Extracto', 'VALOR_Extracto','DESCRIPCION','clave_unica'], axis=1)
    # Caso 4: Salidas en extracto y no en contabilidad
    caso_4 = merged_df[(merged_df['VALOR_Extracto'] < 0) & (merged_df['VALOR_Contabilidad'].isna())]
    caso_4 = caso_4.drop(['FECHA_Contabilidad', 'VALOR_Contabilidad','Concepto Contabilidad','clave_unica'], axis=1)
    # Eliminar la columna de clave temporal
    consolidado.drop('clave_unica', axis=1, inplace=True)

    consolidado = consolidado[['FECHA_Contabilidad', 'VALOR_Contabilidad', 'FECHA_Extracto', 'VALOR_Extracto']]

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:

            # 🔹 Hoja 1: Resultado del join con formato
            # Escribimos el título primero
            merged_df.drop(columns=['clave_unica'], inplace=True)
            merged_df.to_excel(writer, sheet_name='Conciliacion', index=False, startrow=2)
            worksheet = writer.sheets['Conciliacion']
            worksheet.cell(row=1, column=1, value="Resultado de la Conciliación Bancaria").font = Font(bold=True, size=14)

            # Ajustar el ancho de las columnas y formatear fechas (usando openpyxl)
            for col_idx, col in enumerate(worksheet.columns, 1):
                max_length = 0
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

            # 🔹 Hoja 2: Ejemplo de encabezados y tablas
            # Creamos un DataFrame de ejemplo
            caso_1.to_excel(writer, sheet_name='Casos', index=False, startrow=4)
            caso_2.to_excel(writer, sheet_name='Casos', index=False, startrow=4+len(caso_1)+3)
            caso_3.to_excel(writer, sheet_name='Casos', index=False, startrow=4+len(caso_1) + len(caso_2) + 6)
            caso_4.to_excel(writer, sheet_name='Casos', index=False, startrow=4+len(caso_1) + len(caso_2) + len(caso_3) + 9)

            # Accedemos a la hoja 'Ejemplo' para añadir los textos y encabezados
            worksheet_ejemplo = writer.sheets['Casos']
            worksheet_ejemplo.cell(row=1, column=1, value="Formato de Conciliación Bancaria").font = Font(bold=True, size=14)
            worksheet_ejemplo.cell(row=3, column=1, value="Caso 1: Entradas en Contabilidad no en Extracto").font = Font(bold=True)
            worksheet_ejemplo.cell(row=4+len(caso_1)+3, column=1, value="Caso 2: Entradas en Extracto y no en Contabilidad").font = Font(bold=True)
            worksheet_ejemplo.cell(row=4+len(caso_1) + len(caso_2) + 6, column=1, value="Caso 3: Salidas en Contabilidad no en Extracto").font = Font(bold=True)
            worksheet_ejemplo.cell(row=4+len(caso_1) + len(caso_2) + len(caso_3) + 9, column=1, value="Caso 4: Salidas en Extracto y no en Contabilidad").font = Font(bold=True)

            # Ajustar el ancho de las columnas y formatear fechas (usando openpyxl)
            for col_idx, col in enumerate(worksheet_ejemplo.columns, 1):
                max_length = 0
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet_ejemplo.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    # Guardar el archivo Excel en memoria
    output.seek(0)
    return output.read()  # Retornamos los bytes del archivo Excel
